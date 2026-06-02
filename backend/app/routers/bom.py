import io
from typing import List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas, auth
from ..database import get_db

router = APIRouter(prefix="/api/projetos", tags=["bom"])

EDIT_ROLES = (
    models.PerfilEnum.ADMINISTRADOR,
    models.PerfilEnum.ENGENHARIA,
    models.PerfilEnum.PROJETISTA,
)


@router.get("/{projeto_id}/bom", response_model=List[schemas.ProjectBOMResponse])
async def list_boms(
    projeto_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    return (
        db.query(models.ProjectBOM)
        .filter(models.ProjectBOM.projeto_id == projeto_id)
        .order_by(models.ProjectBOM.revisao.desc())
        .all()
    )


@router.post("/{projeto_id}/bom", response_model=schemas.ProjectBOMResponse)
async def create_bom(
    projeto_id: int,
    bom_data: schemas.ProjectBOMCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    projeto = db.query(models.Projeto).filter(models.Projeto.id == projeto_id).first()
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")

    last_bom = (
        db.query(models.ProjectBOM)
        .filter(models.ProjectBOM.projeto_id == projeto_id)
        .order_by(models.ProjectBOM.revisao.desc())
        .first()
    )
    next_rev = (last_bom.revisao + 1) if last_bom else 1

    bom = models.ProjectBOM(
        projeto_id=projeto_id,
        revisao=next_rev,
        descricao=bom_data.descricao,
        status=bom_data.status,
        created_by=current_user.id,
    )
    db.add(bom)
    db.commit()
    db.refresh(bom)
    return bom


@router.get("/{projeto_id}/bom/{bom_id}", response_model=schemas.ProjectBOMResponse)
async def get_bom(
    projeto_id: int,
    bom_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    bom = db.query(models.ProjectBOM).filter(
        models.ProjectBOM.id == bom_id,
        models.ProjectBOM.projeto_id == projeto_id,
    ).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM não encontrada")
    return bom


@router.put("/{projeto_id}/bom/{bom_id}", response_model=schemas.ProjectBOMResponse)
async def update_bom(
    projeto_id: int,
    bom_id: int,
    bom_data: schemas.ProjectBOMCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    bom = db.query(models.ProjectBOM).filter(
        models.ProjectBOM.id == bom_id,
        models.ProjectBOM.projeto_id == projeto_id,
    ).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM não encontrada")

    bom.descricao = bom_data.descricao
    bom.status = bom_data.status
    db.commit()
    db.refresh(bom)
    return bom


@router.post("/{projeto_id}/bom/{bom_id}/items", response_model=schemas.ProjectBOMItemResponse)
async def add_bom_item(
    projeto_id: int,
    bom_id: int,
    item: schemas.ProjectBOMItemCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    bom = db.query(models.ProjectBOM).filter(
        models.ProjectBOM.id == bom_id,
        models.ProjectBOM.projeto_id == projeto_id,
    ).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM não encontrada")

    material = db.query(models.Material).filter(models.Material.id == item.material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material não encontrado")

    db_item = models.ProjectBOMItem(**item.model_dump(), bom_id=bom_id)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item


@router.put(
    "/{projeto_id}/bom/{bom_id}/items/{item_id}",
    response_model=schemas.ProjectBOMItemResponse,
)
async def update_bom_item(
    projeto_id: int,
    bom_id: int,
    item_id: int,
    item: schemas.ProjectBOMItemUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    db_item = db.query(models.ProjectBOMItem).filter(
        models.ProjectBOMItem.id == item_id,
        models.ProjectBOMItem.bom_id == bom_id,
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item não encontrado")

    for key, value in item.model_dump(exclude_unset=True).items():
        setattr(db_item, key, value)

    db.commit()
    db.refresh(db_item)
    return db_item


@router.delete("/{projeto_id}/bom/{bom_id}/items/{item_id}")
async def delete_bom_item(
    projeto_id: int,
    bom_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    db_item = db.query(models.ProjectBOMItem).filter(
        models.ProjectBOMItem.id == item_id,
        models.ProjectBOMItem.bom_id == bom_id,
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item não encontrado")

    db.delete(db_item)
    db.commit()
    return {"message": "Item excluído"}


@router.get("/{projeto_id}/bom/{bom_id}/export")
async def export_bom_excel(
    projeto_id: int,
    bom_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    bom = db.query(models.ProjectBOM).filter(
        models.ProjectBOM.id == bom_id,
        models.ProjectBOM.projeto_id == projeto_id,
    ).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM não encontrada")

    projeto = bom.projeto

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BOM Rev{bom.revisao}"

    from openpyxl.styles import Font, PatternFill, Alignment

    # Header info
    ws["A1"] = "LISTA DE MATERIAIS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Projeto: {projeto.nome}"
    ws["A3"] = f"Código: {projeto.codigo_interno or '-'}"
    ws["A4"] = f"Tipo: {projeto.tipo}"
    ws["A5"] = f"Revisão: {bom.revisao}"
    ws["A6"] = f"Status BOM: {bom.status}"
    ws.append([])

    headers = [
        "Item", "Código", "Família", "Descrição", "Fabricante",
        "Modelo", "Unidade", "Quantidade", "Preço Unit. (R$)",
        "Preço Total (R$)", "Observação",
    ]
    ws.append(headers)

    header_row = ws.max_row
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    for idx, item in enumerate(bom.items, start=1):
        mat = item.material
        subtotal = item.quantidade * item.preco_unitario
        total += subtotal
        ws.append([
            idx,
            mat.codigo_interno or "" if mat else "",
            mat.familia.nome if mat and mat.familia else "",
            mat.descricao if mat else f"Material #{item.material_id}",
            mat.fabricante or "" if mat else "",
            mat.modelo or "" if mat else "",
            mat.unidade if mat else "",
            item.quantidade,
            item.preco_unitario,
            subtotal,
            item.observacao or "",
        ])

    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["", "", "", "", "", "", "", "", "TOTAL", total, ""])
    total_cell = ws.cell(row=total_row, column=10)
    total_cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BOM_{projeto.codigo_interno or projeto.id}_Rev{bom.revisao}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
