import io
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas, auth
from ..database import get_db

router = APIRouter(prefix="/api", tags=["materials"])

EDIT_ROLES = (
    models.PerfilEnum.ADMINISTRADOR,
    models.PerfilEnum.ENGENHARIA,
    models.PerfilEnum.SUPRIMENTOS,
)


# ── Families ──────────────────────────────────────────────────────────────────

@router.get("/familias", response_model=List[schemas.FamiliaResponse])
async def list_familias(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    return db.query(models.FamiliaMaterial).order_by(models.FamiliaMaterial.nome).all()


@router.post("/familias", response_model=schemas.FamiliaResponse)
async def create_familia(
    familia: schemas.FamiliaCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in (models.PerfilEnum.ADMINISTRADOR, models.PerfilEnum.ENGENHARIA):
        raise HTTPException(status_code=403, detail="Permissão insuficiente")
    if db.query(models.FamiliaMaterial).filter(models.FamiliaMaterial.nome == familia.nome).first():
        raise HTTPException(status_code=400, detail="Família já existe")
    db_familia = models.FamiliaMaterial(**familia.model_dump())
    db.add(db_familia)
    db.commit()
    db.refresh(db_familia)
    return db_familia


@router.delete("/familias/{familia_id}")
async def delete_familia(
    familia_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_admin),
):
    db_familia = db.query(models.FamiliaMaterial).filter(models.FamiliaMaterial.id == familia_id).first()
    if not db_familia:
        raise HTTPException(status_code=404, detail="Família não encontrada")
    db.delete(db_familia)
    db.commit()
    return {"message": "Família excluída"}


@router.get("/subfamilias", response_model=List[schemas.SubfamiliaResponse])
async def list_subfamilias(
    familia_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.SubfamiliaMaterial)
    if familia_id:
        query = query.filter(models.SubfamiliaMaterial.familia_id == familia_id)
    return query.order_by(models.SubfamiliaMaterial.nome).all()


@router.post("/subfamilias", response_model=schemas.SubfamiliaResponse)
async def create_subfamilia(
    sub: schemas.SubfamiliaCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in (models.PerfilEnum.ADMINISTRADOR, models.PerfilEnum.ENGENHARIA):
        raise HTTPException(status_code=403, detail="Permissão insuficiente")
    db_sub = models.SubfamiliaMaterial(**sub.model_dump())
    db.add(db_sub)
    db.commit()
    db.refresh(db_sub)
    return db_sub


@router.delete("/subfamilias/{sub_id}")
async def delete_subfamilia(
    sub_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_admin),
):
    db_sub = db.query(models.SubfamiliaMaterial).filter(models.SubfamiliaMaterial.id == sub_id).first()
    if not db_sub:
        raise HTTPException(status_code=404, detail="Subfamília não encontrada")
    db.delete(db_sub)
    db.commit()
    return {"message": "Subfamília excluída"}


# ── Materials ─────────────────────────────────────────────────────────────────

@router.get("/materiais/export/excel")
async def export_materiais_excel(
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    materiais = db.query(models.Material).order_by(models.Material.descricao).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Materiais"

    headers = [
        "Código", "Família", "Subfamília", "Descrição", "Fabricante",
        "Modelo", "Unidade", "Preço Médio (R$)", "Status",
        "Fonte Preço", "Lead Time (dias)", "Observações",
    ]
    ws.append(headers)

    for mat in materiais:
        ws.append([
            mat.codigo_interno or "",
            mat.familia.nome if mat.familia else "",
            mat.subfamilia.nome if mat.subfamilia else "",
            mat.descricao,
            mat.fabricante or "",
            mat.modelo or "",
            mat.unidade,
            mat.preco_medio or 0,
            "Ativo" if mat.status else "Inativo",
            mat.fonte_preco or "",
            mat.lead_time or "",
            mat.observacoes or "",
        ])

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materiais.xlsx"},
    )


@router.get("/materiais", response_model=List[schemas.MaterialResponse])
async def list_materiais(
    search: Optional[str] = None,
    familia_id: Optional[int] = None,
    status: Optional[bool] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.Material)
    if search:
        query = query.filter(
            models.Material.descricao.ilike(f"%{search}%")
            | models.Material.codigo_interno.ilike(f"%{search}%")
            | models.Material.fabricante.ilike(f"%{search}%")
        )
    if familia_id:
        query = query.filter(models.Material.familia_id == familia_id)
    if status is not None:
        query = query.filter(models.Material.status == status)
    return query.order_by(models.Material.descricao).all()


@router.post("/materiais", response_model=schemas.MaterialResponse)
async def create_material(
    material: schemas.MaterialCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    if material.codigo_interno:
        existing = db.query(models.Material).filter(
            models.Material.codigo_interno == material.codigo_interno
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Código interno já existe")

    db_material = models.Material(**material.model_dump(), created_by=current_user.id)
    db.add(db_material)
    db.commit()
    db.refresh(db_material)

    if material.preco_medio and material.preco_medio > 0:
        hist = models.HistoricoPreco(
            material_id=db_material.id,
            preco=material.preco_medio,
            fonte=material.fonte_preco or "Cadastro inicial",
            usuario_id=current_user.id,
        )
        db.add(hist)
        db.commit()

    return db_material


@router.get("/materiais/{material_id}", response_model=schemas.MaterialDetail)
async def get_material(
    material_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    return material


@router.put("/materiais/{material_id}", response_model=schemas.MaterialResponse)
async def update_material(
    material_id: int,
    material: schemas.MaterialUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    db_material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not db_material:
        raise HTTPException(status_code=404, detail="Material não encontrado")

    update_data = material.model_dump(exclude_unset=True)

    if "preco_medio" in update_data and update_data["preco_medio"] != db_material.preco_medio:
        hist = models.HistoricoPreco(
            material_id=material_id,
            preco=update_data["preco_medio"],
            fonte=update_data.get("fonte_preco") or "Atualização manual",
            usuario_id=current_user.id,
        )
        db.add(hist)

    for key, value in update_data.items():
        setattr(db_material, key, value)

    db.commit()
    db.refresh(db_material)
    return db_material


@router.delete("/materiais/{material_id}")
async def delete_material(
    material_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_admin),
):
    db_material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not db_material:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    db.delete(db_material)
    db.commit()
    return {"message": "Material excluído"}


# ── Price History ─────────────────────────────────────────────────────────────

@router.get(
    "/materiais/{material_id}/historico-precos",
    response_model=List[schemas.HistoricoPrecoResponse],
)
async def get_price_history(
    material_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    return (
        db.query(models.HistoricoPreco)
        .filter(models.HistoricoPreco.material_id == material_id)
        .order_by(models.HistoricoPreco.created_at)
        .all()
    )


@router.post(
    "/materiais/{material_id}/historico-precos",
    response_model=schemas.HistoricoPrecoResponse,
)
async def add_price(
    material_id: int,
    preco_data: schemas.HistoricoPrecoCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material não encontrado")

    hist = models.HistoricoPreco(
        **preco_data.model_dump(), material_id=material_id, usuario_id=current_user.id
    )
    db.add(hist)

    material.preco_medio = preco_data.preco
    if preco_data.fonte:
        material.fonte_preco = preco_data.fonte

    db.commit()
    db.refresh(hist)
    return hist


# ── Import Excel ──────────────────────────────────────────────────────────────

@router.post("/materiais/import/excel")
async def import_materiais_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.perfil not in EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    created = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[3]:
            continue

        codigo = str(row[0]).strip() if row[0] else None
        familia_nome = str(row[1]).strip() if row[1] else None
        subfamilia_nome = str(row[2]).strip() if row[2] else None
        descricao = str(row[3]).strip()
        fabricante = str(row[4]).strip() if row[4] else None
        modelo = str(row[5]).strip() if row[5] else None
        unidade = str(row[6]).strip() if row[6] else "un"
        try:
            preco = float(row[7]) if row[7] else 0.0
        except (ValueError, TypeError):
            preco = 0.0

        if not familia_nome:
            errors.append(f"Linha {row_num}: família obrigatória")
            continue

        familia = db.query(models.FamiliaMaterial).filter(
            models.FamiliaMaterial.nome == familia_nome
        ).first()
        if not familia:
            familia = models.FamiliaMaterial(nome=familia_nome)
            db.add(familia)
            db.flush()

        subfamilia_id = None
        if subfamilia_nome:
            subfamilia = db.query(models.SubfamiliaMaterial).filter(
                models.SubfamiliaMaterial.familia_id == familia.id,
                models.SubfamiliaMaterial.nome == subfamilia_nome,
            ).first()
            if not subfamilia:
                subfamilia = models.SubfamiliaMaterial(
                    familia_id=familia.id, nome=subfamilia_nome
                )
                db.add(subfamilia)
                db.flush()
            subfamilia_id = subfamilia.id

        if codigo:
            existing = db.query(models.Material).filter(
                models.Material.codigo_interno == codigo
            ).first()
            if existing:
                errors.append(f"Linha {row_num}: código '{codigo}' já existe (ignorado)")
                continue

        mat = models.Material(
            codigo_interno=codigo,
            familia_id=familia.id,
            subfamilia_id=subfamilia_id,
            descricao=descricao,
            fabricante=fabricante,
            modelo=modelo,
            unidade=unidade,
            preco_medio=preco,
            created_by=current_user.id,
        )
        db.add(mat)
        created += 1

    db.commit()
    return {"created": created, "errors": errors}
